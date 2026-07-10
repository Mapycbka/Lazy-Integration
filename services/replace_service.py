"""Сервис замен по шаблону с использованием openpyxl."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from config.mappings import BSC_PATTERN, BTS_PATTERN, NE_VERSION_NAME_COLUMN


class ReplaceService:
    """Выполняет замены по всей книге и на отдельных листах."""

    def replace_ne_name(self, workbook_path: str | Path, new_name: str, sheet_name: str) -> list[str]:
        """Меняет имя NE на листе NE Version и по всему шаблону.

        Возвращает список старых имен, которые были найдены на листе NE Version.
        """
        workbook = load_workbook(workbook_path, keep_vba=self._should_keep_vba(workbook_path))
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")

        worksheet = workbook[sheet_name]
        header_map = self._get_header_map(worksheet)
        if NE_VERSION_NAME_COLUMN not in header_map:
            raise KeyError(
                f"На листе '{sheet_name}' отсутствует столбец '{NE_VERSION_NAME_COLUMN}'"
            )

        column_index = header_map[NE_VERSION_NAME_COLUMN]
        old_names: list[str] = []
        target_row = 2

        for row in range(target_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            old_value = str(cell.value).strip() if cell.value is not None else ""
            if old_value and old_value not in old_names:
                old_names.append(old_value)

            # На листе NE Version используем только одну строку шаблона.
            if row == target_row:
                cell.value = new_name
            else:
                cell.value = None

        if old_names:
            self._replace_string_values(workbook, old_names, new_name)

        workbook.save(workbook_path)
        workbook.close()
        return old_names

    def replace_bts_software(self, workbook_path: str | Path, new_value: str) -> int:
        """Заменяет версию BTS, сохраняя исходный префикс модели в шаблоне."""
        version_part = self._extract_software_version(new_value, "BTS")
        return self._replace_bts_versions_preserving_prefix(workbook_path, version_part)

    def replace_bsc_software(self, workbook_path: str | Path, new_value: str) -> int:
        """Заменяет все найденные значения BSC-версий на выбранную версию."""
        return self._replace_by_pattern(workbook_path, BSC_PATTERN, new_value)

    def _replace_bts_versions_preserving_prefix(
        self,
        workbook_path: str | Path,
        version_part: str,
    ) -> int:
        """Меняет только версию BTS, не трогая префикс модели."""
        workbook = load_workbook(workbook_path, keep_vba=self._should_keep_vba(workbook_path))
        regex = re.compile(r"(BTS[0-9_ ]+?)(\s*)(V100R\d{3}C\d{2}SPC\d{3})")
        replacements = 0

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        matches = regex.findall(cell.value)
                        if not matches:
                            continue

                        def _replacer(match: re.Match[str]) -> str:
                            prefix = match.group(1)
                            separator = match.group(2) if match.group(2) else " "
                            return f"{prefix}{separator}{version_part}"

                        new_cell_value = regex.sub(_replacer, cell.value)
                        new_cell_value = self._deduplicate_semicolon_values(new_cell_value)
                        if new_cell_value != cell.value:
                            cell.value = new_cell_value
                            replacements += len(matches)

        workbook.save(workbook_path)
        workbook.close()
        return replacements

    def _replace_by_pattern(self, workbook_path: str | Path, pattern: str, new_value: str) -> int:
        workbook = load_workbook(workbook_path, keep_vba=self._should_keep_vba(workbook_path))
        regex = re.compile(pattern)
        replacements = 0

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        matches = regex.findall(cell.value)
                        if matches:
                            new_cell_value = regex.sub(new_value, cell.value)
                            new_cell_value = self._deduplicate_semicolon_values(new_cell_value)
                            if new_cell_value != cell.value:
                                cell.value = new_cell_value
                                replacements += len(matches)

        workbook.save(workbook_path)
        workbook.close()
        return replacements

    @staticmethod
    def _extract_software_version(new_value: str, software_type: str) -> str:
        """Извлекает часть версии без префикса продукта."""
        regex_map = {
            "BTS": r"V100R\d{3}C\d{2}SPC\d{3}",
            "BSC": r"V100R\d{3}C\d{2}SPC\d{3}",
        }
        pattern = regex_map[software_type]
        match = re.search(pattern, new_value)
        if not match:
            raise ValueError(f"Не удалось извлечь версию из значения {software_type}: {new_value}")
        return match.group(0)

    def _replace_string_values(self, workbook, old_values: list[str], new_value: str) -> None:
        """Заменяет найденные старые имена во всех строковых ячейках книги."""
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        updated = cell.value
                        for old_value in old_values:
                            if old_value:
                                updated = updated.replace(old_value, new_value)
                        updated = self._deduplicate_semicolon_values(updated)
                        if updated != cell.value:
                            cell.value = updated

    @staticmethod
    def _get_header_map(worksheet) -> dict[str, int]:
        """Строит карту 'заголовок -> индекс колонки' по первой строке листа."""
        header_map: dict[str, int] = {}
        for index, cell in enumerate(worksheet[1], start=1):
            value = str(cell.value).strip() if cell.value is not None else ""
            if value:
                header_map[value] = index
        return header_map

    @staticmethod
    def _should_keep_vba(workbook_path: str | Path) -> bool:
        """Включает keep_vba только для файлов .xlsm."""
        return Path(workbook_path).suffix.lower() == ".xlsm"

    @staticmethod
    def _deduplicate_semicolon_values(value: str) -> str:
        """Убирает точные дубли значений вида 'A;A' после замен по книге."""
        if ";" not in value:
            return value

        unique_parts: list[str] = []
        seen_normalized: set[str] = set()
        for part in value.split(";"):
            cleaned = part.strip()
            if not cleaned:
                continue
            normalized = re.sub(r"\s+", " ", cleaned).casefold()
            if normalized in seen_normalized:
                continue
            seen_normalized.add(normalized)
            unique_parts.append(cleaned)

        if not unique_parts:
            return ""
        return ";".join(unique_parts)
