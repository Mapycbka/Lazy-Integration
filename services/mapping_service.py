"""Основная бизнес-логика: перенос значений из ДИ/ТС в копию шаблона."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.mappings import (
    ADJNODE_TARGET_COLUMN,
    DI_NE_NAME_COLUMN,
    ENODEB_ID_SOURCE_COLUMN,
    ENODEB_ID_TARGET_COLUMN,
    FOUR_G_DATA_SHEET,
    NE_VERSION_SHEET,
    SCTPLNK_TARGET_COLUMN,
    SITE_CONFIGURATION_SHEET,
    TARGET_TRANSPORT_SHEET,
    TARGET_COLUMN_ALIASES,
    TRANSPORT_COLUMN_MAPPINGS,
    TRANSPORT_SOURCE_SHEET,
)
from services.excel_loader import ExcelLoader
from services.replace_service import ReplaceService
from services.template_manager import TemplateManager


class MappingService:
    """Оркестратор обработки Excel-файлов и заполнения шаблона."""

    def __init__(self, logger: Any) -> None:
        self.logger = logger
        self.excel_loader = ExcelLoader()
        self.template_manager = TemplateManager()
        self.replace_service = ReplaceService()

    def process_files(
        self,
        di_file: str,
        ts_file: str,
        branch: str,
        bbu_type: str,
        template_name: str,
        bts_software: str = "",
        bsc_software: str = "",
        transport_port: str = "",
    ) -> Path:
        """Запускает полный сценарий обработки и возвращает путь к рабочему файлу."""
        self.logger.info("Старт обработки файлов")
        self.logger.info("Выбран филиал: %s, тип BBU: %s", branch, bbu_type)

        site_df = self.excel_loader.load_sheet(di_file, SITE_CONFIGURATION_SHEET)
        ts_df = self.excel_loader.load_sheet(ts_file, TRANSPORT_SOURCE_SHEET)
        four_g_df = self.excel_loader.load_sheet(di_file, FOUR_G_DATA_SHEET)

        ne_name = self.excel_loader.get_first_non_empty_value(
            site_df,
            DI_NE_NAME_COLUMN,
            f"{Path(di_file).name}::{SITE_CONFIGURATION_SHEET}",
        )
        self.logger.info("Новое имя NE: %s", ne_name)

        template_path = self.template_manager.get_template_path(branch, bbu_type, template_name)
        working_copy = self.template_manager.create_working_copy(template_path, ne_name=ne_name)
        self.logger.info("Создана рабочая копия шаблона: %s", working_copy)

        old_names = self.replace_service.replace_ne_name(working_copy, ne_name, NE_VERSION_SHEET)
        if old_names:
            self.logger.info("Заменены старые имена NE: %s", ", ".join(old_names))
        else:
            self.logger.warning(
                "На листе '%s' старое имя NE не найдено, выполнена только запись нового имени",
                NE_VERSION_SHEET,
            )

        if bts_software:
            replacements = self.replace_service.replace_bts_software(working_copy, bts_software)
            self.logger.info("Замена BTS выполнена, найдено совпадений: %s", replacements)

        if bsc_software:
            replacements = self.replace_service.replace_bsc_software(working_copy, bsc_software)
            self.logger.info("Замена BSC выполнена, найдено совпадений: %s", replacements)

        self._fill_transport_sheet(
            working_copy=working_copy,
            ts_df=ts_df,
            site_df=site_df,
            four_g_df=four_g_df,
            transport_port=transport_port,
        )
        self.logger.info("Лист транспортных данных заполнен")

        return working_copy

    def _fill_transport_sheet(
        self,
        working_copy: str | Path,
        ts_df,
        site_df,
        four_g_df,
        transport_port: str,
    ) -> None:
        """Заполняет лист Base Station Transport Data по заголовкам столбцов."""
        transport_row = self.excel_loader.get_first_non_empty_row(
            ts_df,
            list(TRANSPORT_COLUMN_MAPPINGS.keys()),
            f"ТС::{TRANSPORT_SOURCE_SHEET}",
        )
        ne_name = self.excel_loader.get_first_non_empty_value(
            site_df,
            DI_NE_NAME_COLUMN,
            f"ДИ::{SITE_CONFIGURATION_SHEET}",
        )
        enodeb_id = self.excel_loader.get_first_non_empty_value(
            four_g_df,
            ENODEB_ID_SOURCE_COLUMN,
            f"ДИ::{FOUR_G_DATA_SHEET}",
        )

        adjnode_value = self._extract_adjnode(ne_name)
        sctplnk_value = f"{adjnode_value}0"

        workbook = load_workbook(working_copy)
        if TARGET_TRANSPORT_SHEET not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Лист '{TARGET_TRANSPORT_SHEET}' не найден в шаблоне")

        worksheet = workbook[TARGET_TRANSPORT_SHEET]
        header_map, header_row_index = self._get_transport_header_map(worksheet)

        resolved_target_columns = {
            target_name: self._resolve_target_column_name(header_map, target_name)
            for target_name in list(TRANSPORT_COLUMN_MAPPINGS.values()) + [
                ENODEB_ID_TARGET_COLUMN,
                ADJNODE_TARGET_COLUMN,
                SCTPLNK_TARGET_COLUMN,
            ]
        }

        missing_target_columns = [
            original_name for original_name, resolved_name in resolved_target_columns.items()
            if resolved_name is None
        ]
        if missing_target_columns:
            workbook.close()
            missing_str = ", ".join(missing_target_columns)
            raise KeyError(
                f"На листе '{TARGET_TRANSPORT_SHEET}' отсутствуют столбцы: {missing_str}"
            )

        target_row = header_row_index + 1
        for source_column, target_column in TRANSPORT_COLUMN_MAPPINGS.items():
            resolved_target = resolved_target_columns[target_column]
            worksheet.cell(row=target_row, column=header_map[resolved_target]).value = transport_row.get(
                source_column,
                "",
            )

        worksheet.cell(
            row=target_row,
            column=header_map[resolved_target_columns[ENODEB_ID_TARGET_COLUMN]],
        ).value = enodeb_id
        worksheet.cell(
            row=target_row,
            column=header_map[resolved_target_columns[ADJNODE_TARGET_COLUMN]],
        ).value = adjnode_value
        worksheet.cell(
            row=target_row,
            column=header_map[resolved_target_columns[SCTPLNK_TARGET_COLUMN]],
        ).value = sctplnk_value

        if transport_port:
            resolved_transport_port = self._resolve_target_column_name(header_map, "Transport Port")
            if resolved_transport_port:
                worksheet.cell(
                    row=target_row,
                    column=header_map[resolved_transport_port],
                ).value = transport_port
                self.logger.info("В шаблон записан порт транспортной платы: %s", transport_port)
            else:
                self.logger.warning(
                    "Столбец 'Transport Port' отсутствует в шаблоне, выбранный порт %s не записан",
                    transport_port,
                )

        workbook.save(working_copy)
        workbook.close()

    @staticmethod
    def _extract_adjnode(ne_name: str) -> str:
        """Берет последнюю цифровую группу из имени NE и убирает ведущие нули."""
        parts = re.findall(r"\d+", ne_name)
        if not parts:
            raise ValueError(
                "Не удалось вычислить ADJNODE: в имени NE отсутствуют цифровые значения"
            )

        normalized = parts[-1].lstrip("0")
        return normalized or "0"

    @staticmethod
    def _get_header_map(worksheet) -> dict[str, int]:
        """Собирает карту заголовков по первой строке листа."""
        header_map: dict[str, int] = {}
        for index, cell in enumerate(worksheet[1], start=1):
            value = str(cell.value).strip() if cell.value is not None else ""
            if value:
                header_map[value] = index
        return header_map

    @staticmethod
    def _get_transport_header_map(worksheet) -> tuple[dict[str, int], int]:
        """Ищет строку с подстолбцами транспортного листа.

        В некоторых шаблонах заголовки находятся не в первой строке, а во второй,
        потому что первая строка занята группирующими заголовками или merge-ячейками.
        Поэтому проверяем несколько верхних строк и выбираем ту, где найдено
        больше всего ожидаемых имен подстолбцов.
        """
        expected_columns = set(TRANSPORT_COLUMN_MAPPINGS.values()) | {
            ENODEB_ID_TARGET_COLUMN,
            ADJNODE_TARGET_COLUMN,
            SCTPLNK_TARGET_COLUMN,
            "Transport Port",
        }

        best_header_map: dict[str, int] = {}
        best_row_index = 1
        best_score = -1

        for row_index in range(1, min(worksheet.max_row, 5) + 1):
            current_map: dict[str, int] = {}
            for column_index, cell in enumerate(worksheet[row_index], start=1):
                value = str(cell.value).strip() if cell.value is not None else ""
                if value:
                    current_map[value] = column_index

            score = sum(1 for column_name in expected_columns if column_name in current_map)
            if score > best_score:
                best_score = score
                best_header_map = current_map
                best_row_index = row_index

        return best_header_map, best_row_index

    @staticmethod
    def _resolve_target_column_name(
        header_map: dict[str, int],
        target_column_name: str,
    ) -> str | None:
        """Находит фактическое имя столбца в шаблоне с учетом допустимых алиасов."""
        aliases = TARGET_COLUMN_ALIASES.get(target_column_name, [target_column_name])
        for alias in aliases:
            if alias in header_map:
                return alias
        return None
