"""Запись промежуточных сущностей в копию Excel-шаблона."""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.constants import (
    BASE_STATION_TRANSPORT_SHEET,
    SHEET_FIXED_COLUMNS,
    SHEET_DEFAULT_START_ROWS,
    SHEET_HEADER_ALIASES,
    TEMPLATE_SHEET_ALIASES,
    TEMP_DIR,
)
from core.row_cloner import RowCloner
from core.transforms import clone_template_rows
from core.validators import ValidationError
from models.di_models import ProcessedDataBundle
from models.template_models import SheetWriteConfig
from utils.excel_utils import find_best_header_row, resolve_header_name


class TemplateWriter:
    """Создает копию шаблона и заполняет нужные листы по промежуточным моделям."""

    def __init__(self, logger: Any) -> None:
        self.logger = logger
        self.row_cloner = RowCloner()
        TEMP_DIR.mkdir(parents=True, exist_ok=True)

    def create_working_copy(self, template_path: str | Path, ne_name: str) -> Path:
        """Копирует шаблон во временную папку проекта."""
        source = Path(template_path)
        suffix = source.suffix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_name = f"{(ne_name or source.stem).replace('/', '_')}_generated_{timestamp}{suffix}"
        target = TEMP_DIR / target_name
        shutil.copy2(source, target)
        return target

    def write_bundle(
        self,
        working_copy_path: str | Path,
        bundle: ProcessedDataBundle,
    ) -> Path:
        """Заполняет копию шаблона и сохраняет результат."""
        workbook = load_workbook(
            working_copy_path,
            keep_vba=Path(working_copy_path).suffix.lower() == ".xlsm",
        )

        self._write_ne_version(workbook, bundle)
        self._write_auto_deployment(workbook, bundle)
        self._write_base_station_transport_data(workbook, bundle)
        self._write_gsm_cell(workbook, bundle)
        self._write_gtrxgroup(workbook, bundle)
        self._write_lte_cell(workbook, bundle)
        self._write_bbp(workbook, bundle)
        self._write_rruchain(workbook, bundle)
        self._write_rru(workbook, bundle)
        self._write_sector(workbook, bundle)
        self._write_sectoreqm(workbook, bundle)

        workbook.save(working_copy_path)
        workbook.close()
        return self._strip_external_refs(working_copy_path)

    @staticmethod
    def _strip_external_refs(workbook_path: str | Path) -> Path:
        """Удаляет формулы с внешними ссылками [1], которые openpyxl
        иногда ошибочно создаёт при сохранении."""
        import zipfile, io, re

        target = Path(workbook_path)
        data = target.read_bytes()
        z = zipfile.ZipFile(io.BytesIO(data), 'r')
        new_data = io.BytesIO()

        with zipfile.ZipFile(new_data, 'w', zipfile.ZIP_DEFLATED) as out:
            for item in z.infolist():
                content = z.read(item.filename)
                if item.filename.endswith('.xml'):
                    text = content.decode('utf-8', errors='replace')
                    text = re.sub(r'<f[^>]*>.*?\[1\].*?</f>', '', text)
                    content = text.encode('utf-8')
                out.writestr(item, content)

        target.write_bytes(new_data.getvalue())
        return target

    def save_result(self, source_path: str | Path, destination_path: str | Path) -> Path:
        """Сохраняет сформированный результат в выбранное место."""
        try:
            source = Path(source_path)
            destination = Path(destination_path)
            if destination.suffix.lower() != source.suffix.lower():
                destination = destination.with_suffix(source.suffix)
            shutil.copy2(source, destination)
        except Exception as error:
            raise ValidationError(f"Ошибка сохранения результата: {error}") from error
        return destination

    def _write_ne_version(self, workbook, bundle: ProcessedDataBundle) -> None:
        sheet = self._get_template_sheet(workbook, "NE Version")
        config = SheetWriteConfig("NE Version", data_start_row=SHEET_DEFAULT_START_ROWS["NE Version"], single_record_only=True)
        header_map, _ = self._resolve_sheet_headers(sheet, config)
        ne_header = self._resolve_required_header("NE Version", header_map, "NE Name", required=False)
        if ne_header and bundle.ne_name:
            sheet.cell(row=config.data_start_row, column=header_map[ne_header]).value = bundle.ne_name
        else:
            self.logger.warning("На листе 'NE Version' не найден столбец 'NE Name' или NE Name пуст.")

    def _write_gsm_cell(self, workbook, bundle: ProcessedDataBundle) -> None:
        from core.transforms import gsm_freq_band_to_cell_type

        rows = [
            {
                "*BTS Name": bundle.ne_name,
                "*LoCellID": record.ci,
                "*CI": record.ci,
                "BVCI": record.ci,
                "TRX Group ID": record.trx_group_id,
                "*GSM Cell Name": record.cell_name,
                "*LAC": record.lac,
                "NCC": record.ncc,
                "BCC": record.bcc,
                "Routing Area": record.rac,
                "*Frequency of BCCH": record.bcch_frequency,
                "eGBTS Power Type(0.1dBm)": "400",
                "*Cell Type": gsm_freq_band_to_cell_type(record.freq_band),
            }
            for record in bundle.gsm_cells
        ]
        self._write_rows(self._get_template_sheet(workbook, "GSM Cell"), "GSM Cell", rows)

    def _write_base_station_transport_data(self, workbook, bundle: ProcessedDataBundle) -> None:
        if not bundle.transport_data:
            self.logger.warning("Нет данных ТС для заполнения листа 'Base Station Transport Data'.")
            return
        self._write_rows(
            self._get_template_sheet(workbook, BASE_STATION_TRANSPORT_SHEET),
            BASE_STATION_TRANSPORT_SHEET,
            [bundle.transport_data],
        )

    def _write_auto_deployment(self, workbook, bundle: ProcessedDataBundle) -> None:
        if not self._has_template_sheet(workbook, "Auto Deployment"):
            return
        self._write_rows(
            self._get_template_sheet(workbook, "Auto Deployment"),
            "Auto Deployment",
            [{"*Name": bundle.ne_name}],
        )

    def _write_gtrxgroup(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "*BTS Name": bundle.ne_name,
                "*TRX Group ID": record.trx_group_id,
                "*Local Cell ID": record.ci,
                "Sector Equipment ID": record.sector_equipment_id,
            }
            for record in bundle.gsm_cells
        ]
        self._write_rows(self._get_template_sheet(workbook, "GTRXGROUP"), "GTRXGROUP", rows)

    def _write_lte_cell(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "*eNodeB Name": bundle.ne_name,
                "*Cell ID": record.local_cell_id,
                "*LocalCellID": record.local_cell_id,
                "*Cell Name": record.cell_name,
                "*Tracking area code": record.tac,
                "Frequency band": record.frequency_band,
                "Downlink EARFCN": record.downlink_earfcn,
                "Downlink bandwidth": record.downlink_bandwidth,
                "Uplink bandwidth": record.uplink_bandwidth,
                "*Physical cell ID": record.physical_cell_id,
                "Root sequence index": record.root_sequence_index,
                "Reference signal power(0.1dBm)": record.reference_signal_power,
                "PB": record.pb,
                "*Cell transmission and reception mode": record.txrx_mode,
                "*Sector equipment ID": record.sector_equipment_id,
            }
            for record in bundle.lte_cells
        ]
        self._write_rows(self._get_template_sheet(workbook, "LTE Cell"), "LTE Cell", rows)

    def _write_bbp(self, workbook, bundle: ProcessedDataBundle) -> None:
        joined_ports = ", ".join(str(port) for port in bundle.board_ports)
        rows = [{"*Slot No.": slot_number, "Head Port No.": joined_ports} for slot_number in bundle.slot_numbers]
        self._write_rows(self._get_template_sheet(workbook, "BBP(NODE)"), "BBP(NODE)", rows)

    def _write_rruchain(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "*Chain No.": record.chain_no,
                "Head Slot No.": record.head_slot_no,
                "Head Port No.": record.head_port_no,
            }
            for record in bundle.rru_chains
        ]
        self._write_rows(self._get_template_sheet(workbook, "RRUCHAIN(NODE)"), "RRUCHAIN(NODE)", rows)

    def _write_rru(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "Subrack No.": record.subrack_no,
                "*RRU Chain No.": record.chain_no,
                "RRU Name": record.rru_name,
                "RF Unit Working Mode": record.rf_working_mode,
                "Number of RX channels": record.rx_channels,
                "Number of TX channels": record.tx_channels,
            }
            for record in bundle.rrus
        ]
        self._write_rows(self._get_template_sheet(workbook, "RRU(NODE)"), "RRU(NODE)", rows)

    def _write_sector(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "*Sector ID": record.sector_id,
                "Sector Name": record.sector_name,
                "Sector Antenna": record.sector_antenna,
            }
            for record in bundle.sectors
        ]
        self._write_rows(self._get_template_sheet(workbook, "SECTOR(NODE)"), "SECTOR(NODE)", rows)

    def _write_sectoreqm(self, workbook, bundle: ProcessedDataBundle) -> None:
        rows = [
            {
                "*Sector Equipment ID": record.sector_equipment_id,
                "*Sector ID": record.sector_id,
                "Sector Equipment Antenna": record.sector_equipment_antenna,
            }
            for record in bundle.sector_equipments
        ]
        self._write_rows(self._get_template_sheet(workbook, "SECTOREQM(NODE)"), "SECTOREQM(NODE)", rows)

    def _write_rows(self, worksheet, sheet_name: str, rows: list[dict[str, Any]]) -> None:
        if sheet_name not in SHEET_DEFAULT_START_ROWS:
            raise ValidationError(
                f"Для листа '{sheet_name}' не настроена стартовая строка записи в SHEET_DEFAULT_START_ROWS."
            )

        config = SheetWriteConfig(
            sheet_name=sheet_name,
            data_start_row=SHEET_DEFAULT_START_ROWS[sheet_name],
        )
        header_map, header_row = self._resolve_sheet_headers(worksheet, config)
        actual_start_row = max(config.data_start_row, header_row + 1)

        if not rows:
            self.logger.warning("Для листа '%s' нет данных для записи.", sheet_name)
            return

        existing_template_rows = self._count_existing_template_rows(
            worksheet,
            actual_start_row,
            header_map,
        )
        additional_rows_needed = max(0, len(rows) - existing_template_rows)
        if additional_rows_needed > 0:
            clone_template_rows(self.row_cloner, worksheet, actual_start_row, additional_rows_needed)

        for row_offset, row_data in enumerate(rows):
            excel_row = actual_start_row + row_offset
            for logical_header, value in row_data.items():
                fixed_column = SHEET_FIXED_COLUMNS.get(sheet_name, {}).get(logical_header)
                if fixed_column is not None:
                    worksheet.cell(row=excel_row, column=fixed_column).value = value
                    continue

                header_name = self._resolve_required_header(
                    sheet_name,
                    header_map,
                    logical_header,
                    required=False,
                )
                if header_name is None:
                    self.logger.warning(
                        "На листе '%s' отсутствует столбец '%s'. Значение не записано.",
                        sheet_name,
                        logical_header,
                    )
                    continue
                worksheet.cell(row=excel_row, column=header_map[header_name]).value = value

        if existing_template_rows > len(rows):
            rows_to_delete = existing_template_rows - len(rows)
            delete_start_row = actual_start_row + len(rows)
            worksheet.delete_rows(delete_start_row, rows_to_delete)

    @staticmethod
    def _count_existing_template_rows(worksheet, start_row: int, header_map: dict[str, int]) -> int:
        """Оценивает, сколько строк под данные уже заготовлено в шаблоне.

        Берем подряд идущие строки, где заполнена хотя бы одна ячейка в колонках
        рабочей таблицы. Это позволяет не клонировать лишние строки, если шаблон
        уже содержит несколько форматированных заготовок.
        """
        if not header_map:
            return 1

        occupied_rows = 0
        for row_index in range(start_row, worksheet.max_row + 1):
            has_any_value = any(
                worksheet.cell(row=row_index, column=column_index).value not in (None, "")
                for column_index in header_map.values()
            )
            if not has_any_value:
                break
            occupied_rows += 1

        return max(occupied_rows, 1)

    def _resolve_sheet_headers(self, worksheet, config: SheetWriteConfig) -> tuple[dict[str, int], int]:
        aliases = SHEET_HEADER_ALIASES.get(config.sheet_name, {})
        candidate_headers = []
        for header_aliases in aliases.values():
            candidate_headers.extend(header_aliases)
        header_map, header_row = find_best_header_row(
            worksheet,
            candidate_headers,
            config.header_scan_rows,
        )
        return header_map, header_row

    def _resolve_required_header(
        self,
        sheet_name: str,
        header_map: dict[str, int],
        logical_header: str,
        required: bool = True,
    ) -> str | None:
        aliases = SHEET_HEADER_ALIASES.get(sheet_name, {}).get(logical_header, (logical_header,))
        resolved = resolve_header_name(header_map, aliases)
        if resolved is None and required:
            raise ValidationError(
                f"На листе '{sheet_name}' отсутствует обязательный столбец '{logical_header}'."
            )
        return resolved

    @staticmethod
    def _normalize_sheet_name(value: str) -> str:
        normalized = value.casefold().strip()
        for symbol in (" ", "-", "_", "–", "—", ".", ",", "(", ")"):
            normalized = normalized.replace(symbol, "")
        return normalized

    def _get_template_sheet(self, workbook, logical_sheet_name: str):
        aliases = TEMPLATE_SHEET_ALIASES.get(logical_sheet_name, (logical_sheet_name,))
        normalized_lookup = {
            self._normalize_sheet_name(sheet_name): sheet_name
            for sheet_name in workbook.sheetnames
        }
        for alias in aliases:
            normalized_alias = self._normalize_sheet_name(alias)
            if normalized_alias in normalized_lookup:
                return workbook[normalized_lookup[normalized_alias]]
        raise ValidationError(
            f"В шаблоне не найден лист '{logical_sheet_name}' ни по одному из вариантов: {', '.join(aliases)}"
        )

    def _has_template_sheet(self, workbook, logical_sheet_name: str) -> bool:
        aliases = TEMPLATE_SHEET_ALIASES.get(logical_sheet_name, (logical_sheet_name,))
        normalized_sheet_names = {
            self._normalize_sheet_name(sheet_name)
            for sheet_name in workbook.sheetnames
        }
        return any(self._normalize_sheet_name(alias) in normalized_sheet_names for alias in aliases)
