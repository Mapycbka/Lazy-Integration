"""Универсальное клонирование строк шаблона с сохранением форматирования."""

from __future__ import annotations

from copy import copy

from openpyxl.worksheet.worksheet import Worksheet

from utils.excel_utils import copy_cell_style


class RowCloner:
    """Копирует строку-образец вниз вместе со стилями и высотой."""

    def clone_row(self, worksheet: Worksheet, source_row: int, copies_count: int) -> None:
        """Клонирует указанную строку вниз нужное количество раз."""
        if copies_count <= 0:
            return

        for offset in range(1, copies_count + 1):
            target_row = source_row + offset
            worksheet.insert_rows(target_row)
            self._copy_row(worksheet, source_row, target_row)
            self._clone_single_row_merges(worksheet, source_row, target_row)

    def _copy_row(self, worksheet: Worksheet, source_row: int, target_row: int) -> None:
        for column in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=target_row, column=column)
            target_cell.value = source_cell.value
            copy_cell_style(source_cell, target_cell)

        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        worksheet.row_dimensions[target_row].hidden = worksheet.row_dimensions[source_row].hidden
        worksheet.row_dimensions[target_row].outlineLevel = worksheet.row_dimensions[source_row].outlineLevel

    def _clone_single_row_merges(self, worksheet: Worksheet, source_row: int, target_row: int) -> None:
        """Повторяет merge-объединения, если они находятся внутри клонируемой строки."""
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.min_row == source_row and merged_range.max_row == source_row:
                start_column = merged_range.min_col
                end_column = merged_range.max_col
                worksheet.merge_cells(
                    start_row=target_row,
                    start_column=start_column,
                    end_row=target_row,
                    end_column=end_column,
                )

    def clone_template_rows(self, worksheet: Worksheet, source_row: int, copies_count: int) -> None:
        """Публичный алиас с говорящим именем для бизнес-логики записи в шаблон."""
        self.clone_row(worksheet, source_row, copies_count)
