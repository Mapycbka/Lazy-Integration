"""Загрузка Excel-файлов и преобразование листов в словари."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.constants import (
    BSC_NAME_ALIASES,
    GSM_DI_SHEET_ALIASES,
    LTE_DI_SHEET_ALIASES,
    SITE_CONFIGURATION_SHEET,
    SITE_CONFIGURATION_SHEET_ALIASES,
    TRANSPORT_SHEET_ALIASES,
)
from core.validators import ValidationError, validate_di_sheets, validate_template_sheets, validate_ts_sheets
from utils.excel_utils import find_best_header_row
from utils.text_utils import normalize_header, safe_string


class ExcelLoader:
    """Открывает Excel-книги и читает табличные данные листов."""

    def open_di_workbook(self, file_path: str | Path):
        """Открывает файл ДИ для чтения."""
        try:
            workbook = load_workbook(file_path, data_only=False, read_only=False)
        except Exception as error:
            raise ValidationError(f"Невозможно открыть Excel-файл ДИ: {error}") from error

        validate_di_sheets(list(workbook.sheetnames))
        return workbook

    def open_template_workbook(self, file_path: str | Path):
        """Открывает шаблон для записи, сохраняя макросы при наличии."""
        try:
            workbook = load_workbook(file_path, keep_vba=self._should_keep_vba(file_path))
        except Exception as error:
            raise ValidationError(f"Невозможно открыть шаблон Excel: {error}") from error

        validate_template_sheets(list(workbook.sheetnames))
        return workbook

    def open_ts_workbook(self, file_path: str | Path):
        """Открывает файл ТС для чтения."""
        try:
            workbook = load_workbook(file_path, data_only=False, read_only=False)
        except Exception as error:
            raise ValidationError(f"Невозможно открыть Excel-файл ТС: {error}") from error

        validate_ts_sheets(list(workbook.sheetnames))
        return workbook

    def resolve_sheet_name(self, workbook, aliases: tuple[str, ...]) -> str:
        """Ищет реальное имя листа по списку допустимых вариантов."""
        normalized_lookup = {
            self._normalize_sheet_name(sheet_name): sheet_name
            for sheet_name in workbook.sheetnames
        }
        for alias in aliases:
            normalized_alias = self._normalize_sheet_name(alias)
            if normalized_alias in normalized_lookup:
                return normalized_lookup[normalized_alias]

        aliases_text = ", ".join(aliases)
        raise ValidationError(f"Не найден лист ни по одному из вариантов: {aliases_text}")

    def read_sheet_as_dicts(
        self,
        workbook,
        sheet_name: str,
        header_aliases: tuple[str, ...] | None = None,
        max_scan_rows: int = 8,
    ) -> list[dict[str, str]]:
        """Считывает лист в список словарей по найденной строке заголовков."""
        worksheet = workbook[sheet_name]

        if header_aliases:
            header_map, header_row = find_best_header_row(
                worksheet,
                header_aliases,
                max_scan_rows,
            )
            headers = []
            max_column = worksheet.max_column
            for column_index in range(1, max_column + 1):
                header_value = ""
                for header_name, mapped_column_index in header_map.items():
                    if mapped_column_index == column_index:
                        header_value = header_name
                        break
                headers.append(header_value)
        else:
            header_row = 1
            header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
            headers = [safe_string(cell.value) for cell in header_cells]

        if not any(headers):
            raise ValidationError(f"Лист '{sheet_name}' не содержит заголовков.")

        rows: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_dict = {
                headers[index]: safe_string(value)
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            if any(row_dict.values()):
                rows.append(row_dict)
        return rows

    def extract_first_non_empty_value(
        self,
        rows: list[dict[str, str]],
        aliases: tuple[str, ...],
    ) -> str:
        """Ищет первое непустое значение в наборе строк по списку алиасов."""
        for row in rows:
            for alias in aliases:
                value = row.get(alias, "").strip()
                if value:
                    return value
        return ""

    def extract_ne_name(self, workbook) -> str:
        """Пытается получить имя NE из листа 'Конфигурация сайта'."""
        try:
            sheet_name = self.resolve_sheet_name(workbook, SITE_CONFIGURATION_SHEET_ALIASES)
        except ValidationError:
            return ""

        rows = self.read_sheet_as_dicts(workbook, sheet_name)
        return self.extract_first_non_empty_value(
            rows,
            ("Ne Name(name@OSS)", "NE Name", "NodeB Name", "eNodeB Name"),
        )

    def extract_bsc_name(self, workbook) -> str:
        """Пытается получить BSC NAME из листа '2G_BTS_Data'."""
        try:
            sheet_name = self.resolve_sheet_name(workbook, GSM_DI_SHEET_ALIASES)
        except ValidationError:
            return ""

        rows = self.read_sheet_as_dicts(workbook, sheet_name)
        return self.extract_first_non_empty_value(rows, BSC_NAME_ALIASES)

    def resolve_di_sheet_names(self, workbook) -> dict[str, str]:
        """Возвращает реальные имена ключевых листов ДИ."""
        return {
            "2G_BTS_Data": self.resolve_sheet_name(workbook, GSM_DI_SHEET_ALIASES),
            "4G_Data": self.resolve_sheet_name(workbook, LTE_DI_SHEET_ALIASES),
        }

    def extract_first_non_empty_row(
        self,
        rows: list[dict[str, str]],
        aliases: tuple[str, ...],
    ) -> dict[str, str]:
        """Возвращает первую строку, где заполнено хотя бы одно из полей."""
        for row in rows:
            for alias in aliases:
                if row.get(alias, "").strip():
                    return row
        return {}

    @staticmethod
    def _normalize_sheet_name(value: str) -> str:
        """Нормализует имя листа для поиска похожих вариантов написания."""
        normalized = normalize_header(value)
        for symbol in (" ", "-", "_", "–", "—", ".", ","):
            normalized = normalized.replace(symbol, "")
        return normalized

    @staticmethod
    def _should_keep_vba(file_path: str | Path) -> bool:
        """Включает keep_vba только для реально макросных книг."""
        return Path(file_path).suffix.lower() == ".xlsm"
